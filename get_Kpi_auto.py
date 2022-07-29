import pandas as pd
import numpy as np
import xlsxwriter
import json
from googleapiclient.errors import HttpError
import os
from Google import Create_Service
import codecs
from Google import *
from openpyxl import load_workbook
import datetime

# считывание таблицы___________________________________________________________________________________________________
a = batch_get_values("1UcaizxVvwXzCKTEUvV8SsnpwaNWf_TPuYfuB3Z7v-pM", 'ПЭС АЛМ ежедн. СВОД')
first = a['valueRanges'][0]['values'][5]
second = a['valueRanges'][0]['values'][6]
third = a['valueRanges'][0]['values'][7]
for i in range(len(first)):
    try:
        n = second[i]
    except:
        second.insert(i, '')
    try:
        m = third[i]
    except:
        third.insert(i, '')

# парсинг заголовков таблицы___________________________________________________________________________________________
new_col_names = []
for e in range(len(first)):
    if first[e] != '' and second[e] != '':
        new_col_names.append(first[e]+'_'+second[e]+third[e])
    elif first[e] == '' and second[e] != '':
        first[e] = str(first[e-1])
        new_col_names.append(first[e]+'_'+second[e]+third[e])
    elif second[e] == '' and first[e] != '':
        new_col_names.append(first[e])
    else:
        if first[e] == '' and second[e] == '':
            new_col_names.append(first[e - 2] + '_' + second[e - 2] + '_' + third[e])
        else:
            new_col_names.append(first[e-1]+'_'+second[e-1]+'_'+third[e])

new_col_names.append("ID")
val_range = a['valueRanges'][0]['values'][9:]

for i in range(len(val_range)):
    for j in range(0, 48):
        try:
            n = val_range[i][j]
        except:
            val_range[i].insert(j, '')
    val_range[i].insert(48, i+1)

new_val_range=[] # новый массив без пустых строк
for i in range(len(val_range)):
    if val_range[i][0] != '' and val_range[i][2] != '' and val_range[i][8] != '':
        new_val_range.append(val_range[i])

data = dict()
for i in range(len(new_col_names)):
    data[new_col_names[i]] = list(map(lambda v: v[i], new_val_range))
print(new_col_names)

# запись в json________________________________________________________________________________________________________
df = pd.DataFrame(data=data)
df.set_index('ID', inplace=True)
df_new = df['ГРНЗ'].groupby([df['ДАТА']]).apply(list)

result_ex = df_new.to_json(orient='index')
parsed_ex = json.loads(result_ex)
json_result_ex = json.dumps(parsed_ex, ensure_ascii=False, indent=4)

with open('weight_result_exel.json', "w", encoding='utf-8') as outfile:
    outfile.write(json_result_ex)
with open('weight_result_exel.json') as f:
    file_content = f.read()
    GRNS_by_data = json.loads(file_content)

# Рассчет Kpi-ГСМ______________________________________________________________________________________________________
Days_all = list(GRNS_by_data.keys())
GRNS_all = list(GRNS_by_data.items())
Cp_GSM_dict = {}
for i in range(len(GRNS_all)): # счетчик дней
    Cp_GSM_day = {}
    for j in range(len(GRNS_all[i][1])): # счетчик транспорта
        GRNS_by_day = list(filter(lambda x: GRNS_all[i][1][j] in x and Days_all[i] in x, new_val_range))
        MSK = 0
        all_trip = 0
        if (GRNS_by_day[0][30] != "0,0") and (GRNS_by_day[0][30] != '') and (GRNS_by_day[0][43] != ''):
            rashod = GRNS_by_day[0][43].replace(",", ".")
            probeg = GRNS_by_day[0][30].replace(",", ".")
            tonns = GRNS_by_day[0][36].replace(",", ".")
            Cp_GSM = (float(rashod) / float(probeg)) * 100
            MSK = int(GRNS_by_day[0][32])
            all_trip = int(GRNS_by_day[0][31])
            tonns = float(tonns)
            Cp_GSM_day [GRNS_all[i][1][j]] = Cp_GSM, GRNS_by_day[0][19], MSK, all_trip, tonns # значение Ср_ГСМ, норма расхода ГСМ, счетчик посещений МСК, счетчк всех рейсов, тонны
        else:
            continue
    Cp_GSM_dict[Days_all[i]] = Cp_GSM_day

# Вытаскиваем все ГРНЗ номера__________________________________________________________________________________________
df_GRNS = df['ДАТА'].groupby([df['ГРНЗ']]).apply(list)
GRNS = list(df_GRNS.keys())

Cp_GSM_month = {}
Cp_GSM_sum = ''
for i in range (len(GRNS)):
    Cp_GSM_sum = 0
    days_count = 0 # счетчик рабочих дней
    MSK_count = 0 # счетчик посещений МСК
    Trips_count = 0 # счетчик всех рейсов
    Weight = 0
    for j in range(len(Days_all)):
        day = Days_all[j]
        GRNS_number = GRNS[i]
        if Cp_GSM_dict[day].get(GRNS_number, False):# проверка на наличие ключа в словаре
            Cp_GSM_one_day = Cp_GSM_dict[day][GRNS_number][0]
            Cp_GSM_sum = Cp_GSM_sum + float(Cp_GSM_one_day) # складываем Ср_ГСМ за все дни
            days_count = days_count + 1
            Norma = Cp_GSM_dict[day][GRNS_number][1]
            MSK_count = MSK_count + Cp_GSM_dict[day][GRNS_number][2]
            Trips_count = Trips_count + Cp_GSM_dict[day][GRNS_number][3]
            Weight = Weight + Cp_GSM_dict[day][GRNS_number][4]
        else:
            continue
    if days_count > 0:
        Cp_GSM_sum = Cp_GSM_sum / days_count
        RD = MSK_count / days_count
        Cpp = Trips_count / days_count
        if Trips_count:
            Cpt = Weight/Trips_count
        else:
            Cpt = 0
    else:
        Cp_GSM_sum = 0
        RD = 0
        Cpp = 0
    Cp_GSM_month[GRNS[i]] = Cp_GSM_sum, Norma, RD, Cpp, Cpt

k = 3
KPI_dict = {}
for i in range (len(GRNS)):
    GRNS_number = GRNS[i]
    RD = Cp_GSM_month[GRNS_number][2]
    Cpp = Cp_GSM_month[GRNS_number][3]
    Cp_GSM = float(Cp_GSM_month[GRNS_number][0])
    Ngsm = Cp_GSM_month[GRNS_number][1].replace(",", ".")
    Ngsm = float(Ngsm)
    if (Cp_GSM-Ngsm)>3:
        KPI_dict[GRNS_number] = 0, Cp_GSM_month[GRNS_number][2], RD, Cpp
    else:
        KPI = (Ngsm/k) + 1 - (1/k*Cp_GSM)
        KPI_final = KPI ** (0.5)
        # {Номер_ГРНЗ: Kpi_гсм, РД, Срр}
        KPI_dict [GRNS_number] = KPI_final, RD, Cpp

# Получение норм_______________________________________________________________________________________________________
norms_path = "Exel/Constans/Нормы ТС.xlsx"
wb = load_workbook(norms_path)
sheet = wb.get_sheet_by_name('NORMA')
max_row = sheet.max_row

GRNS_norms = {}
for row in range (3, max_row):
    if sheet[row][0].value in GRNS:
        GRNS_norms [sheet[row][0].value] = {'norm_trips': sheet[row][1].value, 'norm_weight_winter': sheet[row][2].value,
                                            'norm_weight_summer': sheet[row][3].value, 'norm_GSM_winter': sheet[row][4].value,
                                            'norm_GSM_summer': sheet[row][5].value}
    else:
        continue

# Рассчет Kpi-Рейсы____________________________________________________________________________________________________
Kpi_trip_dict = {}
for i in range(len(GRNS)):
    GRNS_number = GRNS[i]
    if GRNS_norms.get(GRNS_number, False):
        Kpi = 2 ** (KPI_dict[GRNS_number][2] - GRNS_norms[GRNS_number]['norm_trips'])
    else:
        Kpi = 0
    Kpi_trip_dict[GRNS_number] = Kpi

# Рассчет Kpi-Тоннаж___________________________________________________________________________________________________
Kpi_tonn_dict = {}
for i in range(len(GRNS)):
    GRNS_number = GRNS[i]
    if GRNS_norms.get(GRNS_number, False):
        now = datetime.datetime.now()
        if (now.month > 3) and (now.month > 10):
            Nt = GRNS_norms[GRNS_number]['norm_weight_summer']
        else:
            Nt = GRNS_norms[GRNS_number]['norm_weight_winter']
        Cpt = int(Cp_GSM_month[GRNS_number][4]) * 1000
        if Cpt > Nt:
            Kpi = 1
        else:
            Kpi = ((-1/Nt ** 2) * (Cpt ** 2)) + (2 * Cpt/Nt)
    else:
        Kpi = 0
    Kpi_tonn_dict[GRNS_number] = Kpi

# Итоговая премия машины_______________________________________________________________________________________________
Pst_dict = {}
for i in range(len(GRNS)):
    GRNS_number = GRNS[i]
    S = 160000
    Kpi_tablet = 1
    Pst = S * KPI_dict[GRNS_number][1] * (0.5 * Kpi_trip_dict[GRNS_number] + 0.15 * Kpi_tonn_dict[GRNS_number] + 0.15 * KPI_dict[GRNS_number][0] + 0.2 * Kpi_tablet)
    Pst_dict[GRNS_number] = Pst
print(Pst_dict)

with open("Json/Premiya_auto.json", "w", encoding="utf-8") as file:
    json.dump(Pst_dict, file, sort_keys=False, indent=4, ensure_ascii=False)

with open("Norms/Norms_auto.json", "w", encoding="utf-8") as file:
    json.dump(GRNS_norms, file, sort_keys=False, indent=4, ensure_ascii=False)


