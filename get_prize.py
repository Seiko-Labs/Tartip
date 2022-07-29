import xlsxwriter
import json
import codecs

from openpyxl.styles import Font

from Google import *
import openpyxl
from openpyxl.writer.excel import save_workbook


# Данные о транспорте и водителях__________________________________________________________________
with open('Json/Premiya_auto.json') as f:
    file_content = f.read()
    GRNS_all = json.loads(file_content)
print(GRNS_all)

fileObj = codecs.open("Json/Premiya_drivers.json", "r", "utf_8_sig")
text = fileObj.read()
Drivers_all = eval(text)
print(Drivers_all)

# Данные грузчиков__________________________________________________________________
fileObj = codecs.open("Json/Loaders_GRNS.json", "r", "utf_8_sig")
text = fileObj.read()
Loaders_all = eval(text)

GRNS = list(GRNS_all.keys()) #все ГРНЗ
FIO = list(Drivers_all.keys()) #все ФИО водителей
FIO_loaders = list(Loaders_all.keys()) #все ФИО грузчиков
FIO_loaders.remove('')

#Сумма водителей_____________________________________________________________________________________________
Drivers_prize = {}

for i in range(len(GRNS)):
    GRNS_number = GRNS[i]
    Prize_all_drivers = 0
    for j in range(len(FIO)):
        FIO_number = FIO[j]
        for k in range(len(Drivers_all[FIO_number]['Pst_list'])):
            if Drivers_all[FIO_number]['Pst_list'][k].get(GRNS_number, False):
                Prize_all_drivers = Prize_all_drivers + Drivers_all[FIO_number]['Pst_list'][k][GRNS_number]
    Drivers_prize[GRNS_number] = Prize_all_drivers

#Итоговая премия водителей_____________________________________________________________________________________________
Prize_dict = {}

for i in range(len(FIO)):
    FIO_number = FIO[i]
    Prize_list = []
    for k in range(len(Drivers_all[FIO_number]['Pst_list'])):
        Prize_GRNS = {}
        GRNS_temp = list(Drivers_all[FIO_number]['Pst_list'][k].keys())
        GRNS_number = GRNS_temp[0]
        if Drivers_prize.get(GRNS_number, False):
            Prize = GRNS_all[GRNS_number] * Drivers_all[FIO_number]['Pst_list'][k][GRNS_number] / Drivers_prize[GRNS_number]
            Prize_GRNS[GRNS_number] = Prize
        else:
            Prize_GRNS[GRNS_number] = 0
        Prize_list.append(Prize_GRNS)
    Prize_dict[FIO_number] = Prize_list


#Итоговая премия грузчиков_____________________________________________________________________________________________
Loaders_prize = {}
for i in range(len(FIO_loaders)):
    FIO_number = FIO_loaders[i]
    Prize_load_list = []
    if FIO_number != '':
        for k in range(len(Loaders_all[FIO_number]['GRNS_list'])):
            Prize_GRNS = {}
            GRNS_number = Loaders_all[FIO_number]['GRNS_list'][k]
            days = Loaders_all[FIO_number]['Work_days']
            if days > 10:
                Prize_GRNS = (GRNS_all[GRNS_number] * 0.65)/days
            else:
                Prize_GRNS = 0
            Prize_load_list.append(Prize_GRNS)
        Loaders_prize[FIO_number] = {'job': Loaders_all[FIO_number]['job'], 'District': Loaders_all[FIO_number]['District'], 'Pst_list': Prize_load_list}


now = datetime.datetime.now()
today = now.strftime("%d-%m-%Y")
month = now.strftime("%m-%Y")

# открываем новый файл на запись

exel_name = 'Salary/' + 'Kpi_prize_' + str(month) + '.xlsx'

try:
    workbook = openpyxl.load_workbook(exel_name)
    worksheet = workbook.create_sheet(str(today))

    worksheet.column_dimensions['A'].width = 6
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15

    worksheet['A1'] = '№'
    worksheet['A1'].font = Font(bold=True)
    worksheet['B1'] = 'Район'
    worksheet['B1'].font = Font(bold=True)
    worksheet['C1'] = 'ФИО'
    worksheet['C1'].font = Font(bold=True)
    worksheet['D1'] = 'Должность'
    worksheet['D1'].font = Font(bold=True)
    worksheet['E1'] = 'ГРНЗ'
    worksheet['E1'].font = Font(bold=True)
    worksheet['F1'] = 'Премия за СТ'
    worksheet['F1'].font = Font(bold=True)
    worksheet['G1'] = 'ИТОГ'
    worksheet['G1'].font = Font(bold=True)

    line_count = 2
# Водители__________________________
    for i in range(len(FIO)):
        ID_spot = 'A' + str(line_count)
        worksheet[ID_spot] = i + 1

        FIO_number = FIO[i]
        FIO_spot = 'C' + str(line_count)
        worksheet[FIO_spot] = FIO_number

        District = Drivers_all[FIO_number]['District']
        District_spot = 'B' + str(line_count)
        worksheet[District_spot] = District

        job = Drivers_all[FIO_number]['job']
        job_spot = 'D' + str(line_count)
        worksheet[job_spot] = job

        Itog = 0
        Itog_spot = 'G' + str(line_count)

        for k in range(len(Prize_dict[FIO_number])):
            GRNS_temp = list(Drivers_all[FIO_number]['Pst_list'][k].keys())
            GRNS_number = GRNS_temp[0]
            GRNS_spot = 'E' + str(line_count)
            worksheet[GRNS_spot] = GRNS_number

            Prize_spot = 'F' + str(line_count)
            Prize = Prize_dict[FIO_number][k][GRNS_number]
            worksheet[Prize_spot] = Prize
            Itog += Prize

            line_count += 1
        Last_id = i
        Last_line = line_count
        worksheet[Itog_spot] = Itog
    line_count = line_count + 1



    for i in range(len(FIO_loaders)):
        ID_spot = 'A' + str(line_count)
        worksheet[ID_spot] = i + 1 + Last_id

        FIO_number = FIO_loaders[i]
        FIO_spot = 'C' + str(line_count)
        worksheet[FIO_spot] = FIO_number

        District = Loaders_prize[FIO_number]['District']
        District_spot = 'B' + str(line_count)
        worksheet[District_spot] = District

        job = Loaders_prize[FIO_number]['job']
        job_spot = 'D' + str(line_count)
        worksheet[job_spot] = job

        Itog = 0
        Itog_spot = 'G' + str(line_count)

        for k in range(len(Loaders_prize[FIO_number]['Pst_list'])):
            GRNS_number = Loaders_prize[FIO_number]['Pst_list'][k]
            GRNS_spot = 'E' + str(line_count)
            worksheet[GRNS_spot] = GRNS_number

            Prize_spot = 'F' + str(line_count)
            Prize = Loaders_prize[FIO_number]['Pst_list'][k]
            worksheet[Prize_spot] = Prize
            Itog += Prize
            line_count += 1

        worksheet[Itog_spot] = Itog
    save_workbook(workbook, exel_name)
    workbook.close()

except:
    workbook = xlsxwriter.Workbook(exel_name)
    worksheet = workbook.add_worksheet(str(today))  # создаем лист

    worksheet.set_column(0, 0, 6) # №
    worksheet.set_column(2, 2, 30) # ФИО
    worksheet.set_column(3, 3, 20) # Должность
    worksheet.set_column(5, 6, 15) # Премия за СТ и ИТОГ

    bold = workbook.add_format({'bold': True})
    worksheet.write('A1', '№', bold)
    worksheet.write('B1', 'Район', bold)
    worksheet.write('C1', 'ФИО', bold)
    worksheet.write('D1', 'Должность', bold)
    worksheet.write('E1', 'ГРНЗ', bold)
    worksheet.write('F1', 'Премия за СТ', bold)
    worksheet.write('G1', 'ИТОГ', bold)

#Водители___________________________________________________________________________________________________________
    line_count = 2
    for i in range(len(FIO)):
        ID_spot = 'A' + str(line_count)
        worksheet.write(ID_spot, i + 1)

        FIO_number = FIO[i]
        FIO_spot = 'C' + str(line_count)
        worksheet.write(FIO_spot, FIO_number)

        District = Drivers_all[FIO_number]['District']
        District_spot = 'B' + str(line_count)
        worksheet.write(District_spot, District)

        job = Drivers_all[FIO_number]['job']
        job_spot = 'D' + str(line_count)
        worksheet.write(job_spot, job)

        Itog = 0
        Itog_spot = 'G' + str(line_count)

        for k in range(len(Prize_dict[FIO_number])):
            GRNS_temp = list(Drivers_all[FIO_number]['Pst_list'][k].keys())
            GRNS_number = GRNS_temp[0]
            GRNS_spot = 'E' + str(line_count)
            print(GRNS_number)
            print(GRNS_spot)
            worksheet.write(GRNS_spot, GRNS_number)

            Prize_spot = 'F' + str(line_count)
            Prize = Prize_dict[FIO_number][k][GRNS_number]
            worksheet.write(Prize_spot, Prize)
            Itog += Prize
            line_count += 1
        Last_id = i
        Last_line = line_count
        worksheet.write(Itog_spot, Itog)
    workbook.close()

# Грузчики___________________________________________________________________________________________________________
    workbook = openpyxl.load_workbook(exel_name)
    worksheet = workbook[str(today)]
    line_count = line_count + 1
    for i in range(len(FIO_loaders)):
        ID_spot = 'A' + str(line_count)
        worksheet[ID_spot] = i + 1 + Last_id

        FIO_number = FIO_loaders[i]
        FIO_spot = 'C' + str(line_count)
        worksheet[FIO_spot] = FIO_number

        District = Loaders_prize[FIO_number]['District']
        District_spot = 'B' + str(line_count)
        worksheet[District_spot] = District

        job = Loaders_prize[FIO_number]['job']
        job_spot = 'D' + str(line_count)
        worksheet[job_spot] = job

        Itog = 0
        Itog_spot = 'G' + str(line_count)

        for k in range(len(Loaders_prize[FIO_number]['Pst_list'])):
            GRNS_number = Loaders_prize[FIO_number]['Pst_list'][k]
            GRNS_spot = 'E' + str(line_count)
            worksheet[GRNS_spot] = GRNS_number

            Prize_spot = 'F' + str(line_count)
            Prize = Loaders_prize[FIO_number]['Pst_list'][k]
            worksheet[Prize_spot] = Prize
            Itog += Prize
            line_count += 1

        worksheet[Itog_spot] = Itog
    save_workbook(workbook, exel_name)


