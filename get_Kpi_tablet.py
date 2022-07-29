import json
import pathlib
from openpyxl import load_workbook

ff_path = "Exel/Покрытие+вывоза+фотоматериалом_01_03_2022_01_03_2022.xlsx" #фф по документации
f_path = "Exel/Сводная+статистика+по+выходу+водителей+с+планшетом_01_03_2022_01.xlsx" #ф по документации
ttn_path = "Exel/Корректность+ввода+данных+водителями_01_03_2022_01_03_2022.xlsx" #ттн по документации
wb = load_workbook(ff_path)
sheet = wb.active
max_row = sheet.max_row
max_column = sheet.max_column

# считаем ФФ___________________________________________________________________________________________________________
FF_dict = {}
for row in range (6, max_row):
     try:
          fact_task = int(sheet[row][2].value)
          norm_task = int(sheet[row][3].value)
          if norm_task>0:
               ff_num = fact_task/norm_task
          else:
               ff_num = 0
          FF_dict[sheet[row][1].value] = ff_num
     except:
          continue

# считаем Ф____________________________________________________________________________________________________________
wb_f = load_workbook(f_path)
sheet_f = wb_f.active
max_row = sheet_f.max_row
max_column = sheet_f.max_column
F_dict = {}
for row in range (6, max_row):
     try:
          tablet_naryad = int(sheet_f[row][2].value)
          norm_naryad = int(sheet_f[row][1].value)
          if norm_naryad>0:
               f_num = tablet_naryad/norm_naryad
          else:
               f_num = 0
          F_dict[sheet_f[row][0].value] = f_num
     except:
          continue

# считаем ТТН__________________________________________________________________________________________________________
wb_ttn = load_workbook(ttn_path)
sheet_ttn = wb_ttn.active
max_row = sheet_ttn.max_row
max_column = sheet_ttn.max_column
TTN_dict = {}
for row in range (4, max_row):
     try:
          District = sheet_ttn[row][3].value.split()
          TTN_num = float(sheet_ttn[row][2].value)
          TTN_dict [sheet_ttn[row][1].value] = {'TTN': TTN_num, 'District': District[-1]}
     except:
          continue

# считаем Kpi планшета_________________________________________________________________________________________________
All_FIOs = list(FF_dict.keys())
Kpi_tablet_dict = {}
for fio in range (len(All_FIOs)):
     try:
          TTN = TTN_dict[All_FIOs[fio]]['TTN'] / 100
          KPI_tablet = FF_dict[All_FIOs[fio]] * FF_dict[All_FIOs[fio]] * TTN

          Kpi_tablet_dict[All_FIOs[fio]] = {'KPI': KPI_tablet}
     except:
          Kpi_tablet_dict[All_FIOs[fio]] = {'KPI': 0}

print(Kpi_tablet_dict)
with open("Json/Kpi_tablet.json", "w", encoding="utf-8") as file:
    json.dump(Kpi_tablet_dict, file, sort_keys=False, indent=4, ensure_ascii=False)
